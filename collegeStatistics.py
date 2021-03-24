'''
Install "mysqlclient" package to enable python connectivity to mysql (that you already installed).
Now do the following tasks from a python click script - not using mysql client or workbench,
you can use them to verify what you did. Implement the following commands - take arguments as appropriate

createdb -> command that creates a database and the appropriate tables (students and marks)
            and foreign key relationship between marks and students tables
dropdb -> drop the database
importdata -> import the data in the students.xlsx in assignment 2 and marks.xlsx that you generated in assignment 3 into the tables.
collegestats -> print out a console report showing number of students per college
               and the min, max and average marks of students from the college. [college acronym, student count, min, avg, max]
'''

import click
import MySQLdb
from MySQLdb import Error
from openpyxl import *
import warnings
warnings.filterwarnings("ignore")

@click.group()
def statistics():
    pass

@statistics.command()
def createdb():
    """ Connect to MySQL database and create Srudents and Marks tables"""

    try:
        conn = MySQLdb.connect(host="localhost", user="root", passwd="raviprince57")
        cur1 = conn.cursor()  # prepare a cursor object using cursor() method
        cur2 = conn.cursor()

        """ Creating a  Schema or DataBase """
        cur1.execute("DROP SCHEMA IF EXISTS statistics")  ### Drop the DATABASE if it already exist ###

        query = "CREATE SCHEMA `statistics`;"
        cur1.execute(query)

        """ Creating students and marks tables"""

        query = """CREATE TABLE statistics.STUDENTS (
                    NAME  CHAR(50) NOT NULL,
                    COLLEGE CHAR(50) NOT NULL,
                    EMAILID CHAR(50) NOT NULL,
                    DBNAMES CHAR(50) NOT NULL )"""
        cur1.execute(query)

        query = """CREATE TABLE statistics.MARKS (
                        STUDENT  CHAR(40) NOT NULL,
                        TRANFORM INT NOT NULL,
                        FROM_CUSTOM_BASE26 INT NOT NULL,
                        GET_PIG_LATIN INT NOT NULL,
                        TOP_CHARS INT NOT NULL,
                        TOTAL INT NOT NULL )"""
        cur2.execute(query)
        conn.close()
        print "Successfully created database and the appropriate tables (students and marks)"
    except Error as e:
        print(e)

@statistics.command()
def dropdb():
    " Drops the Database"
    try:
        conn = MySQLdb.connect(host="localhost", user="root", passwd="raviprince57")
        query = "DROP SCHEMA IF EXISTS statistics"
        cursor = conn.cursor()
        cursor.execute(query)
        conn.close()
        print "Database droped successfully"
    except Error as e:
        print(e)

@statistics.command()
@click.argument("studentsfile",nargs=1)
@click.argument("marksfile",nargs=1)
def importdata(studentsfile,marksfile):
    """Imports the data into the tables from excel files"""

    # wb = load_workbook(excel_file, read_only=True)
    # ws = wb['Colleges']
    #
    # table_data = map(lambda row : {
    #     'name': row[0].value,
    #     'location' : row[1].value,
    #     'acronym' : row[2].value,
    #     'contact' : row[3].value }
    #     ,ws[2:ws.max_row])

    try:
        conn = MySQLdb.connect(host="localhost", user="root", passwd="raviprince57", db="statistics")
        cur1 = conn.cursor()
        cur2 = conn.cursor()

        ## importing data from students table ##
        wb = load_workbook(studentsfile, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(row_offset=1):
            row_data = [col.value.encode(encoding="ascii").strip() for col in row]
            query = """INSERT INTO STUDENTS(NAME,COLLEGE, EMAILID,DBNAMES) VALUES """ + str(tuple(row_data))
            cur1.execute(query)

        conn.commit()  ####  commit the changes in the database ####

        ## importing data from marks table ##
        wb = load_workbook(marksfile, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(row_offset=1):
            row_data = [col.value.encode(encoding="ascii") if col.data_type == 's' else int(col.value) for col in row]
            query = """INSERT INTO MARKS
                              (STUDENT, TRANFORM, FROM_CUSTOM_BASE26, GET_PIG_LATIN, TOP_CHARS, TOTAL)
                               VALUES """ + str(tuple(row_data))
            cur2.execute(query)

        conn.commit()  ####  commit the changes in the database ####
        conn.close()
        print "Data imported successfully"
    except Error as e:
        print(e)

@statistics.command()
def collegestats():
    """ Generates the college report """
    " Connecting to the data base and creating cursor objects"
    conn = MySQLdb.connect(host="localhost", user="root", passwd="raviprince57", db="statistics")
    cur1 = conn.cursor()
    cur2 = conn.cursor()

    # query = "SELECT DISTINCT COLLEGE FROM STUDENTS"
    # cur1.execute(query)
    # distinct_colleges = cur1.fetchall()
    #
    # for college in distinct_colleges:
    #     query = "SELECT COUNT(TOTAL),AVG(TOTAL),MAX(TOTAL),MIN(TOTAL) FROM MARKS WHERE STUDENT LIKE 'ol2016_" + college[0] + "%'"
    #     cur2.execute(query)
    #     result = cur2.fetchall()[0]
    #     print college[0], int(result[0]), float(result[1]), int(result[2]), int(result[3])
    query = "SELECT COUNT(COLLEGE),COLLEGE FROM STUDENTS GROUP BY COLLEGE"
    cur1.execute(query)
    students_count = { tup[1]:tup[0] for tup in cur1.fetchall()} ## collegename : count

    query = "SELECT STUDENTS.COLLEGE,MAX(MARKS.TOTAL),MIN(MARKS.TOTAL),AVG(MARKS.TOTAL) FROM MARKS , STUDENTS " \
            "WHERE MARKS.STUDENT LIKE CONCAT(CONCAT('%',STUDENTS.COLLEGE),'%') GROUP BY STUDENTS.COLLEGE"
    cur1.execute(query)
    result = cur1.fetchall()
    for college_report in result:
        count = students_count[college_report[0]]
        print college_report[0],count,college_report[1],college_report[2],float(college_report[3])

    conn.close()

if __name__=='__main__':
    statistics()