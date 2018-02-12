"""  Assignment2 """
'''
This assignment is about reading and writing excel files through python. Refer to the guidelines in @33.
This builds on assignment 1 (on using click).
Here's the high level spec, you have to figure out all the details and get this done:

Create a script called copyexcel.py which uses openpyxl and click.
It copies all data from one excel file to another and transforms them according to specified criteria.
If the input has more than one worksheet, then all of them are copied.
Script usage is as follows:

copyexcel.py --capitalize --preservestyles <source_excel> <dest_excel>
--capitalize  -> boolean flag. default false. If specified, all string data will be capitalized during copy into destination
--preservestyles -> boolean falg.  default false. If specified, even the cell styles will be copied, else only data.
source_excel -> argument that specifies the input file
dest_excel -> argument that specifies the output file. It already exists, prompt if user wants to overwrite. If yes, overwrite it.

You can use this as the sample data file for testing: students.xlsx
'''

import click
from openpyxl import *
from os.path import exists
from copy import copy

def copy_the_content(src_filepath,dest_filepath,capitalize,preservestyles):
    src_wb = load_workbook(src_filepath)
    dest_wb = Workbook()
    dest_wb.remove_sheet(dest_wb.active)
    for sheet in src_wb:
        dest_curr_ws = dest_wb.create_sheet(sheet.title)
        alph = 65
        num=1
        for col in sheet.iter_cols():
            dest_curr_ws.column_dimensions[chr(alph)].width =  sheet.column_dimensions[chr(alph)].width
            alph+=1
            for cell in col:
                if capitalize:
                   dest_curr_ws[cell.coordinate] = cell.value.capitalize() if cell.data_type=='s' else cell.value
                if preservestyles:
                    dest_curr_ws[cell.coordinate].style = copy(cell.style)
                    dest_curr_ws[cell.coordinate].font = copy(cell.font)
                    dest_curr_ws[cell.coordinate].border = copy(cell.border)
                    dest_curr_ws[cell.coordinate].alignment = copy(cell.alignment)
                    dest_curr_ws[cell.coordinate].fill = copy(cell.fill)
                    dest_curr_ws[cell.coordinate].protection = copy(cell.protection)


    dest_wb.save(dest_filepath)

@click.command()
@click.option("--capitalize",is_flag=True,help="If specified, all string data will be capitalized during copy into destination.")
@click.option("--preservestyles",is_flag=True,help="If specified, even the cell styles will be copied, else only data.")
@click.argument("source_excel",nargs=1)
@click.argument("dest_excel",nargs=1)

def copying_excel_data(capitalize,preservestyles,source_excel,dest_excel):
    if exists(dest_excel):
        overwrite = raw_input("The destination file already exists, do you want to overwrite it : Y/N ?\n")
        if overwrite == 'Y':
           copy_the_content(source_excel,dest_excel,capitalize,preservestyles)
    else:
        copy_the_content(source_excel,dest_excel,capitalize,preservestyles)


if __name__ == "__main__":
    copying_excel_data()


